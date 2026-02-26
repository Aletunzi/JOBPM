import io
import math
from datetime import datetime, timezone, timedelta
from typing import Optional

from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import select, func, and_, or_, delete
from sqlalchemy.ext.asyncio import AsyncSession

from api.auth import require_api_key
from api.cache import cache_get, cache_set, cache_clear
from api.database import get_db
from api.models import Company, Job
from api.schemas import CompanyOut, CompaniesResponse, CompanyPatch


router = APIRouter(prefix="/api/companies", tags=["companies"])

DEFAULT_LIMIT = 50
MAX_LIMIT = 200


async def _company_to_out(c: Company, active_jobs: int) -> CompanyOut:
    return CompanyOut(
        id=c.id,
        name=c.name,
        website_url=getattr(c, "website_url", None),
        career_url=c.career_url,
        career_url_source=getattr(c, "career_url_source", "auto"),
        tier=c.tier,
        size=c.size,
        vertical=c.vertical,
        geo_primary=c.geo_primary,
        is_enabled=c.is_enabled,
        last_scraped=c.last_scraped,
        scrape_status=getattr(c, "scrape_status", None),
        active_jobs=active_jobs,
    )


@router.get("", response_model=CompaniesResponse)
async def list_companies(
    page: int = Query(1, ge=1),
    limit: int = Query(DEFAULT_LIMIT, ge=1, le=MAX_LIMIT),
    search: Optional[str] = Query(None, description="Filter by company name (substring)"),
    vertical: Optional[str] = Query(None, description="Comma-separated verticals"),
    geo: Optional[str] = Query(None, description="Comma-separated geo_primary values"),
    tier: Optional[str] = Query(None, description="Comma-separated tiers: 1,2,3"),
    enabled: Optional[bool] = Query(None, description="Filter by is_enabled"),
    has_url: Optional[bool] = Query(None, description="Filter to companies with/without career_url"),
    status: Optional[str] = Query(None, description="Filter by scrape_status (OK, EMPTY, HTTP_ERROR, SPA_DETECTED)"),
    sort: Optional[str] = Query(None, description="Sort order: name_asc, name_desc, last_scraped_asc, last_scraped_desc"),
    db: AsyncSession = Depends(get_db),
    _key: str = Depends(require_api_key),
):
    cache_key = f"companies:{page}:{limit}:{search}:{vertical}:{geo}:{tier}:{enabled}:{has_url}:{status}:{sort}"
    cached = cache_get(cache_key)
    if cached:
        return cached

    conditions = []

    if search and search.strip():
        conditions.append(Company.name.ilike(f"%{search.strip()}%"))

    if vertical:
        v_list = [v.strip().lower() for v in vertical.split(",") if v.strip()]
        if v_list:
            conditions.append(Company.vertical.in_(v_list))

    if geo:
        g_list = [g.strip().upper() for g in geo.split(",") if g.strip()]
        if g_list:
            conditions.append(Company.geo_primary.in_(g_list))

    if tier:
        t_list = [int(t.strip()) for t in tier.split(",") if t.strip().isdigit()]
        if t_list:
            conditions.append(Company.tier.in_(t_list))

    if enabled is not None:
        conditions.append(Company.is_enabled == enabled)

    if has_url is True:
        conditions.append(Company.career_url.is_not(None))
    elif has_url is False:
        conditions.append(Company.career_url.is_(None))

    if status:
        conditions.append(Company.scrape_status == status.upper())

    where_clause = and_(*conditions) if conditions else True

    # Total count
    total = await db.scalar(
        select(func.count()).select_from(Company).where(where_clause)
    )
    total = total or 0
    pages = math.ceil(total / limit) if total > 0 else 1
    offset = (page - 1) * limit

    # Active jobs count per company via correlated subquery
    active_jobs_subq = (
        select(func.count(Job.id))
        .where(and_(Job.company_id == Company.id, Job.is_active == True))
        .correlate(Company)
        .scalar_subquery()
    )

    if sort == "name_asc":
        order_by = [Company.name.asc()]
    elif sort == "name_desc":
        order_by = [Company.name.desc()]
    elif sort == "last_scraped_asc":
        order_by = [Company.last_scraped.asc().nulls_last()]
    elif sort == "last_scraped_desc":
        order_by = [Company.last_scraped.desc().nulls_last()]
    else:
        order_by = [Company.tier.asc(), Company.name.asc()]

    stmt = (
        select(Company, active_jobs_subq.label("active_jobs"))
        .where(where_clause)
        .order_by(*order_by)
        .offset(offset)
        .limit(limit)
    )

    result = await db.execute(stmt)
    rows = result.all()

    items = [
        await _company_to_out(company, active_jobs or 0)
        for company, active_jobs in rows
    ]

    response = CompaniesResponse(items=items, total=total, page=page, pages=pages)
    cache_set(cache_key, response)
    return response


@router.get("/export")
async def export_companies_excel(
    db: AsyncSession = Depends(get_db),
    _key: str = Depends(require_api_key),
):
    """Export all companies as an Excel (.xlsx) file."""
    active_jobs_subq = (
        select(func.count(Job.id))
        .where(and_(Job.company_id == Company.id, Job.is_active == True))
        .correlate(Company)
        .scalar_subquery()
    )

    stmt = (
        select(Company, active_jobs_subq.label("active_jobs"))
        .order_by(Company.tier.asc(), Company.name.asc())
    )
    result = await db.execute(stmt)
    rows = result.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Companies"
    ws.append(["Name", "Vertical", "Geo", "Tier", "Size", "Website URL", "Career URL", "Source",
               "Status", "Last Scraped", "Active Jobs", "Enabled", "Action"])

    for company, active_jobs in rows:
        ws.append([
            company.name,
            company.vertical or "",
            company.geo_primary or "",
            company.tier,
            company.size or "",
            getattr(company, "website_url", "") or "",
            company.career_url or "",
            getattr(company, "career_url_source", "auto"),
            getattr(company, "scrape_status", "") or "",
            company.last_scraped.strftime("%Y-%m-%d %H:%M") if company.last_scraped else "",
            active_jobs or 0,
            "Yes" if company.is_enabled else "No",
            "",  # Action: leave blank or write DELETE to remove on import
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=fursa_companies.xlsx"},
    )


@router.post("/import")
async def import_companies_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _key: str = Depends(require_api_key),
):
    """
    Import an Excel file (same format as export) to bulk-update companies.

    Editable columns (matched by Name):
      - Website URL  → website_url
      - Career URL   → career_url
      - Enabled      → is_enabled  (Yes / No)
      - Action       → write DELETE to remove the company and its jobs
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx / .xls files are accepted.")

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not parse the file. Make sure it is a valid Excel file.")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="The file is empty.")

    # Build column index from header row
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    def col(name: str) -> Optional[int]:
        try:
            return header.index(name)
        except ValueError:
            return None

    idx_name       = col("Name")
    idx_website    = col("Website URL")
    idx_career     = col("Career URL")
    idx_enabled    = col("Enabled")
    idx_action     = col("Action")

    if idx_name is None:
        raise HTTPException(status_code=400, detail="'Name' column not found in the file.")

    stats = {"updated": 0, "deleted": 0, "not_found": [], "errors": []}

    for row in rows[1:]:
        name = str(row[idx_name]).strip() if row[idx_name] is not None else ""
        if not name:
            continue

        result = await db.execute(select(Company).where(Company.name == name))
        company = result.scalar_one_or_none()
        if company is None:
            stats["not_found"].append(name)
            continue

        # Check for DELETE action first
        action = str(row[idx_action]).strip().upper() if idx_action is not None and row[idx_action] else ""
        if action == "DELETE":
            await db.execute(delete(Job).where(Job.company_id == company.id))
            await db.delete(company)
            stats["deleted"] += 1
            continue

        # Apply editable fields
        changed = False
        if idx_website is not None and row[idx_website] is not None:
            website = str(row[idx_website]).strip()
            if website != (company.website_url or ""):
                company.website_url = website or None
                changed = True
        if idx_career is not None and row[idx_career] is not None:
            career = str(row[idx_career]).strip()
            if career != (company.career_url or ""):
                company.career_url = career or None
                company.career_url_source = "manual" if career else "auto"
                if not career:
                    company.page_hash = None
                changed = True
        if idx_enabled is not None and row[idx_enabled] is not None:
            enabled_val = str(row[idx_enabled]).strip().lower()
            enabled = enabled_val in ("yes", "true", "1")
            if enabled != company.is_enabled:
                company.is_enabled = enabled
                changed = True

        if changed:
            stats["updated"] += 1

    await db.commit()
    cache_clear()
    return stats


@router.patch("/{company_id}", response_model=CompanyOut)
async def patch_company(
    company_id: UUID,
    patch: CompanyPatch,
    db: AsyncSession = Depends(get_db),
    _key: str = Depends(require_api_key),
):
    """Update mutable fields of a company (website_url, career_url, is_enabled)."""
    from api.models import Company

    company = await db.get(Company, company_id)
    if company is None:
        raise HTTPException(status_code=404, detail="Company not found")

    fields_set = patch.model_fields_set

    if "website_url" in fields_set:
        company.website_url = patch.website_url

    if "career_url" in fields_set:
        company.career_url = patch.career_url
        # Track provenance: manual if setting a URL, auto if clearing (allow rediscovery)
        company.career_url_source = "manual" if patch.career_url else "auto"
        if not patch.career_url:
            company.page_hash = None

    if "is_enabled" in fields_set:
        company.is_enabled = patch.is_enabled

    await db.commit()
    await db.refresh(company)
    cache_clear()

    active_jobs = await db.scalar(
        select(func.count(Job.id)).where(
            and_(Job.company_id == company.id, Job.is_active == True)
        )
    )
    return await _company_to_out(company, active_jobs or 0)


@router.delete("/{company_id}", status_code=204)
async def delete_company(
    company_id: UUID,
    db: AsyncSession = Depends(get_db),
    _key: str = Depends(require_api_key),
):
    """Delete a company and all its jobs."""
    company = await db.get(Company, company_id)
    if company is None:
        raise HTTPException(status_code=404, detail="Company not found")
    await db.execute(delete(Job).where(Job.company_id == company_id))
    await db.delete(company)
    await db.commit()
    cache_clear()
